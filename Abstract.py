import os
import sqlite3
import openpyxl
from PyQt6 import QtCore, QtGui, QtWidgets, QtPrintSupport
from PyQt6.QtCore import QFileInfo
from PyQt6.QtGui import QPainter, QPageLayout, QPageSize, QFont
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
from PyQt6.QtWidgets import QMessageBox, QTableWidgetSelectionRange, QFileDialog
from openpyxl import Workbook
from openpyxl.styles import Font


class Abstract_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1011, 788)
        font = QtGui.QFont()
        font.setFamily("Helvetica") # Set font
        font.setPointSize(12)
        Dialog.setFont(font)

        self.verticalLayout = QtWidgets.QVBoxLayout(Dialog)
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem)

        self.pushButton_sort = QtWidgets.QPushButton(parent=Dialog)
        # icon with relative path
        icon1 = QtGui.QIcon()
        image_path_to_icon1 = os.path.join(os.path.dirname(__file__), "images", "sort-alphabet.png")
        icon1.addPixmap(QtGui.QPixmap(image_path_to_icon1), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        self.pushButton_sort.setIcon(icon1)
        self.pushButton_sort.setObjectName("pushButton_sort")

        # Set focus policy to NoFocus
        self.pushButton_sort.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)

        # Connect signal
        self.pushButton_sort.clicked.connect(self.sort)
        self.horizontalLayout_2.addWidget(self.pushButton_sort)

        # icon with relative path
        self.pushButton_exportToExcel = QtWidgets.QPushButton(parent=Dialog)
        icon2 = QtGui.QIcon()
        image_path_to_icon2 = os.path.join(os.path.dirname(__file__), "images", "table-excel.png")
        icon2.addPixmap(QtGui.QPixmap(image_path_to_icon2), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        self.pushButton_exportToExcel.setIcon(icon2)
        self.pushButton_exportToExcel.setObjectName("pushButton_exportToExcel")

        # Set focus policy to NoFocus
        self.pushButton_exportToExcel.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)

        # Connect Signal
        self.pushButton_exportToExcel.clicked.connect(self.export_to_excel)
        self.horizontalLayout_2.addWidget(self.pushButton_exportToExcel)

        # icon with relative path
        self.pushButton_exportToPdf = QtWidgets.QPushButton(parent=Dialog)
        icon3 = QtGui.QIcon()
        image_path_to_icon3 = os.path.join(os.path.dirname(__file__), "images", "pdf.png")
        icon3.addPixmap(QtGui.QPixmap(image_path_to_icon3), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        self.pushButton_exportToPdf.setIcon(icon3)
        self.pushButton_exportToPdf.setObjectName("pushButton_exportToPdf")

        # Set focus policy to NoFocus
        self.pushButton_exportToPdf.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)

        # Connect signal
        self.pushButton_exportToPdf.clicked.connect(self.export_pdf)

        self.horizontalLayout_2.addWidget(self.pushButton_exportToPdf)

        # icon with relative path
        self.pushButton_printPreview = QtWidgets.QPushButton(parent=Dialog)
        icon4 = QtGui.QIcon()
        image_path_to_icon4 = os.path.join(os.path.dirname(__file__), "images", "printprev.png")
        icon4.addPixmap(QtGui.QPixmap(image_path_to_icon4), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        self.pushButton_printPreview.setIcon(icon4)
        self.pushButton_printPreview.setObjectName("pushButton_printPreview")

        # Set focus policy to NoFocus
        self.pushButton_printPreview.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)

        self.horizontalLayout_2.addWidget(self.pushButton_printPreview)

        # icon with relative path
        self.pushButton_print = QtWidgets.QPushButton(parent=Dialog)
        icon5 = QtGui.QIcon()
        image_path_to_icon5 = os.path.join(os.path.dirname(__file__), "images", "print.png")
        icon5.addPixmap(QtGui.QPixmap(image_path_to_icon5), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        self.pushButton_print.setIcon(icon5)
        self.pushButton_print.setObjectName("pushButton_print")

        # Set focus policy to NoFocus
        self.pushButton_print.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)

        # Connect signal
        self.pushButton_print.clicked.connect(self.print)

        self.horizontalLayout_2.addWidget(self.pushButton_print)

        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem1)
        self.verticalLayout.addLayout(self.horizontalLayout_2)


        self.tableWidget_takeOff = QtWidgets.QTableWidget(parent=Dialog)
        self.tableWidget_takeOff.setObjectName("tableWidget_takeOff")
        self.tableWidget_takeOff.setColumnCount(9)
        self.tableWidget_takeOff.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_takeOff.setHorizontalHeaderItem(8, item)
        self.verticalLayout.addWidget(self.tableWidget_takeOff)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout.addItem(spacerItem2)

        # icon with relative path
        self.pushButton_cancel = QtWidgets.QPushButton(parent=Dialog)
        icon6 = QtGui.QIcon()
        image_path_to_icon6 = os.path.join(os.path.dirname(__file__), "images", "x-circle-duotone.png")
        icon6.addPixmap(QtGui.QPixmap(image_path_to_icon6), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        self.pushButton_cancel.setIcon(icon6)
        self.pushButton_cancel.setObjectName("pushButton_cancel")

        # Set focus policy to NoFocus
        self.pushButton_cancel.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)

        self.horizontalLayout.addWidget(self.pushButton_cancel)

        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout.addItem(spacerItem3)
        self.verticalLayout.addLayout(self.horizontalLayout)

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Abstract"))
        self.pushButton_sort.setText(_translate("Dialog", "Sort"))
        self.pushButton_exportToExcel.setText(_translate("Dialog", "Export to Excel"))
        self.pushButton_exportToPdf.setText(_translate("Dialog", "Export to Pdf"))
        self.pushButton_printPreview.setText(_translate("Dialog", "Print Preview"))
        self.pushButton_print.setText(_translate("Dialog", "Print"))
        item = self.tableWidget_takeOff.horizontalHeaderItem(0)
        item.setText(_translate("Dialog", "code"))
        self.tableWidget_takeOff.setColumnWidth(0, 70)

        item = self.tableWidget_takeOff.horizontalHeaderItem(1)
        item.setText(_translate("Dialog", "trade"))
        self.tableWidget_takeOff.setColumnWidth(1, 40)

        item = self.tableWidget_takeOff.horizontalHeaderItem(2)
        item.setText(_translate("Dialog", "desc"))
        self.tableWidget_takeOff.setColumnWidth(2, 350)

        item = self.tableWidget_takeOff.horizontalHeaderItem(3)
        item.setText(_translate("Dialog", "ref"))
        self.tableWidget_takeOff.setColumnWidth(3, 70)

        item = self.tableWidget_takeOff.horizontalHeaderItem(4)
        item.setText(_translate("Dialog", "times"))
        self.tableWidget_takeOff.setColumnWidth(4, 70)

        item = self.tableWidget_takeOff.horizontalHeaderItem(5)
        item.setText(_translate("Dialog", "dims"))
        self.tableWidget_takeOff.setColumnWidth(5, 60)

        item = self.tableWidget_takeOff.horizontalHeaderItem(6)
        item.setText(_translate("Dialog", "square"))
        self.tableWidget_takeOff.setColumnWidth(6, 90)

        item = self.tableWidget_takeOff.horizontalHeaderItem(7)
        item.setText(_translate("Dialog", "unit"))
        self.tableWidget_takeOff.setColumnWidth(7, 40)

        item = self.tableWidget_takeOff.horizontalHeaderItem(8)
        item.setText(_translate("Dialog", "sign post"))
        self.tableWidget_takeOff.setColumnWidth(8, 170)

        self.pushButton_cancel.setText(_translate("Dialog", "Cancel"))

        # Run load_tabel_data method
        self.load_table_data()

        # apply table style sheet
        self.apply_table_stylesheet()

    def load_table_data(self):
        # Connect to the SQLite database
        conn = sqlite3.connect('data.db')
        cursor = conn.cursor()

        # Get the list of table names from the database
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()

        # Extract table names from the fetched data and store them in a list
        table_list = [table[0] for table in tables if table[0] != 'sqlite_sequence']  # Exclude the 'sqlite_sequence'

        # print(table_list)

        # Initialize an empty list to store all the retrieved data
        all_data = []

        # Iterate over each table in table_list
        for table_name in table_list:
            # Retrieve the data from the table
            cursor.execute(f'SELECT * FROM {table_name}')
            data = cursor.fetchall()

            # Append the retrieved data to the all_data list
            all_data.extend(data)

        # print(all_data)

        # Check if the all_data list is empty
        if not all_data:
            return

        # Set the number of rows and columns in the QTableWidget
        self.tableWidget_takeOff.setRowCount(len(all_data))
        self.tableWidget_takeOff.setColumnCount(9)  # Column count set to 9

        # Populate the QTableWidget with the retrieved data
        for row_num, row_data in enumerate(all_data):
            square_value = 0.0  # Initialize square_value
            sign_post_value = None  # Initialize sign_post_value

            for col_num, col_data in enumerate(row_data):
                if col_num == 0:  # Exclude ID column
                    continue

                item = QtWidgets.QTableWidgetItem()

                if col_num == 7:  # "square" column is at index 7
                    try:
                        square_value = float(col_data)
                        formatted_value = '{:,.2f}'.format(square_value)
                    except ValueError:
                        formatted_value = str(col_data)
                else:
                    formatted_value = str(col_data)

                item.setText(formatted_value)
                self.tableWidget_takeOff.setItem(row_num, col_num - 1, item)  # Adjust column index

                if col_num == 9:  # "sign_post" column is at index 9
                    sign_post_value = str(col_data)  # Store the value of the sign_post column

            # Check if square value is negative and set row color to red
            if square_value < 0:
                for col in range(self.tableWidget_takeOff.columnCount()):
                    cell_item = self.tableWidget_takeOff.item(row_num, col)
                    cell_item.setForeground(QtGui.QColor('red'))

            # Check if sign_post_value == "SUM", set row foreground to blue
            if sign_post_value == "SUM":
                for col in range(self.tableWidget_takeOff.columnCount()):
                    cell_item = self.tableWidget_takeOff.item(row_num, col)
                    cell_item.setForeground(QtGui.QColor('blue'))

            # Check if sign_post_value == "SUM", set row foreground to blue
            if sign_post_value == "TONNAGE":
                for col in range(self.tableWidget_takeOff.columnCount()):
                    cell_item = self.tableWidget_takeOff.item(row_num, col)
                    cell_item.setForeground(QtGui.QColor('green'))

        # Freeze the table contents
        self.tableWidget_takeOff.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)

        # Close the database connection
        conn.close()

    def sort(self):
        # Set the sort in ascending order (A-Z)
        sort_order = QtCore.Qt.SortOrder.AscendingOrder

        # Sort the table by the "trade" column
        self.tableWidget_takeOff.sortItems(1, sort_order)

    def export_to_excel(self):
        # Create a workbook and select the active sheet
        workbook = Workbook()
        sheet = workbook.active

        # Export table headers
        headers = []
        for col in range(self.tableWidget_takeOff.columnCount()):
            header = self.tableWidget_takeOff.horizontalHeaderItem(col).text()
            headers.append(header)
            sheet.cell(row=1, column=col + 1).value = header

        # Export table data with formatting
        for row in range(self.tableWidget_takeOff.rowCount()):
            for col in range(self.tableWidget_takeOff.columnCount()):
                # Get the data and foreground color from each cell in the QTableWidget
                cell_item = self.tableWidget_takeOff.item(row, col)
                cell_data = cell_item.text()
                foreground_color = cell_item.foreground().color()

                # Write the cell data to the corresponding cell in the Excel sheet
                sheet.cell(row=row + 2, column=col + 1).value = cell_data

                # Apply font color formatting based on the foreground color
                if foreground_color == QtGui.QColor('red'):
                    font = Font(color="FFFF0000")  # Red
                elif foreground_color == QtGui.QColor('blue'):
                    font = Font(color="FF0000FF")  # Blue
                elif foreground_color == QtGui.QColor('green'):
                    font = Font(color="FF00FF00")  # Green
                else:
                    font = None

                # Apply font color formatting to the corresponding cell in the Excel sheet
                if font:
                    sheet.cell(row=row + 2, column=col + 1).font = font

                # Add vertical borders to each column
                border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                right=openpyxl.styles.Side(style='thin'))
                sheet.cell(row=row + 2, column=col + 1).border = border

        # Create an "exports" directory if it doesn't exist
        export_dir = os.path.join(os.getcwd(), 'exports')
        os.makedirs(export_dir, exist_ok=True)

        # Save the workbook to an Excel file in the "exports" directory
        export_path = os.path.join(export_dir, 'takeOff_sheet.xlsx')
        workbook.save(export_path)

        # Show a message box with the location of the saved file
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle("Export Successful")
        msg_box.setText(f"Take-Offs exported to: {export_path}")

        # Add icon with relative path
        icon_path = os.path.join(os.path.dirname(__file__), "images", "exclamation-circle.png")
        icon_pixmap = QtGui.QPixmap(icon_path)
        msg_box.setIconPixmap(icon_pixmap)

        msg_box.exec()

    def apply_table_stylesheet(self):
        self.tableWidget_takeOff.setStyleSheet("QTableView::item { border-right: 1px solid black; }")
        self.tableWidget_takeOff.setFont(QFont("Helvetica", 12))

    def print(self):
        pass

    def export_pdf(self):
        pass